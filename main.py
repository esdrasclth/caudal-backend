from fastapi import FastAPI, Query
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from supabase import create_client
import pandas as pd
import io
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from dotenv import load_dotenv
import os
from datetime import datetime

load_dotenv()

app = FastAPI(title="Caudal API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

supabase = create_client(os.getenv("SUPABASE_URL"), os.getenv("SUPABASE_KEY"))

def get_transacciones(user_id: str, mes: str):
    inicio = f"{mes}-01"
    año = int(mes.split("-")[0])
    m = int(mes.split("-")[1])
    import calendar
    ultimo_dia = calendar.monthrange(año, m)[1]
    fin = f"{mes}-{ultimo_dia:02d}"

    try:
        result = supabase.table("transactions") \
    .select("*, categories(nombre, icono), wallets!transactions_wallet_id_fkey(nombre)") \
    .eq("user_id", user_id) \
    .gte("fecha", inicio) \
    .lte("fecha", fin) \
    .order("fecha", desc=False) \
    .execute()

        return result.data if result.data else []
    except Exception as e:
        print(f"Error obteniendo transacciones: {e}")
        return []

@app.get("/")
def root():
    return {"mensaje": "Caudal API funcionando 💧"}

@app.get("/debug")
def debug(user_id: str = Query(...), mes: str = Query(...)):
    transacciones = get_transacciones(user_id, mes)
    return {
        "user_id": user_id,
        "mes": mes,
        "total": len(transacciones),
        "primera": transacciones[0] if transacciones else None
    }

@app.get("/exportar/excel")
def exportar_excel(user_id: str = Query(...), mes: str = Query(...)):
    try:
        transacciones = get_transacciones(user_id, mes)
        print(f"Transacciones encontradas: {len(transacciones)}")
        if transacciones:
            print(f"Primera transacción: {transacciones[0]}")

        if not transacciones:
            from fastapi.responses import JSONResponse
            return JSONResponse({"error": "No hay transacciones"}, status_code=404)

        filas = []
        for t in transacciones:
            cats = t.get("categories") or {}
            walls = t.get("wallets") or {}
            filas.append({
                "Fecha": t.get("fecha", ""),
                "Descripción": t.get("descripcion", "") or "",
                "Categoría": cats.get("nombre", "") if isinstance(cats, dict) else "",
                "Cartera": walls.get("nombre", "") if isinstance(walls, dict) else "",
                "Tipo": "Ingreso" if t.get("tipo") == "ingreso" else "Gasto",
                "Monto (HNL)": float(t.get("monto", 0))
            })

        df = pd.DataFrame(filas)
        total_ingresos = df[df["Tipo"] == "Ingreso"]["Monto (HNL)"].sum()
        total_gastos = df[df["Tipo"] == "Gasto"]["Monto (HNL)"].sum()
        saldo = total_ingresos - total_gastos

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Transacciones", index=False)
            ws = writer.sheets["Transacciones"]

            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for col in ws.iter_cols(min_row=1, max_row=1):
                for cell in col:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

            for col in ws.columns:
                max_width = max(len(str(cell.value or "")) for cell in col) + 4
                ws.column_dimensions[col[0].column_letter].width = min(max_width, 40)

            green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
            red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
            for row in ws.iter_rows(min_row=2):
                tipo_cell = row[4]
                fill = green_fill if tipo_cell.value == "Ingreso" else red_fill
                for cell in row:
                    cell.fill = fill

            resumen_data = {
                "Concepto": ["Total Ingresos", "Total Gastos", "Saldo Neto"],
                "Monto (HNL)": [total_ingresos, total_gastos, saldo]
            }
            df_resumen = pd.DataFrame(resumen_data)
            df_resumen.to_excel(writer, sheet_name="Resumen", index=False)
            ws_resumen = writer.sheets["Resumen"]
            for col in ws_resumen.iter_cols(min_row=1, max_row=1):
                for cell in col:
                    cell.fill = header_fill
                    cell.font = header_font

        buffer.seek(0)
        nombre_mes = datetime.strptime(mes, "%Y-%m").strftime("%B_%Y")

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=Caudal_{nombre_mes}.xlsx"}
        )
    except Exception as e:
        print(f"Error exportando Excel: {e}")
        import traceback
        traceback.print_exc()
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": str(e)}, status_code=500)

@app.get("/exportar/pdf")
def exportar_pdf(user_id: str = Query(...), mes: str = Query(...)):
    transacciones = get_transacciones(user_id, mes)
    nombre_mes = datetime.strptime(mes, "%Y-%m").strftime("%B %Y").capitalize()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=0.75*inch,
        leftMargin=0.75*inch,
        topMargin=0.75*inch,
        bottomMargin=0.75*inch
    )

    styles = getSampleStyleSheet()
    elements = []

    # Título
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Title"],
        fontSize=22,
        textColor=colors.HexColor("#0D9488"),
        spaceAfter=4,
        alignment=TA_CENTER
    )
    subtitle_style = ParagraphStyle(
        "Subtitle",
        parent=styles["Normal"],
        fontSize=12,
        textColor=colors.HexColor("#64748B"),
        spaceAfter=20,
        alignment=TA_CENTER
    )

    elements.append(Paragraph("💧 Caudal — Finanzas Personales", title_style))
    elements.append(Paragraph(f"Reporte de {nombre_mes}", subtitle_style))

    # Calcular resumen
    total_ingresos = sum(float(t.get("monto", 0)) for t in transacciones if t.get("tipo") == "ingreso")
    total_gastos = sum(float(t.get("monto", 0)) for t in transacciones if t.get("tipo") == "gasto")
    saldo = total_ingresos - total_gastos

    def fmt(n):
        return f"L {n:,.2f}"

    # Cards de resumen
    resumen_data = [
        ["💰 Ingresos", "💸 Gastos", "📊 Saldo Neto"],
        [fmt(total_ingresos), fmt(total_gastos), fmt(saldo)]
    ]

    resumen_table = Table(resumen_data, colWidths=[2.2*inch, 2.2*inch, 2.2*inch])
    resumen_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#DCFCE7")),
        ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#FEE2E2")),
        ("BACKGROUND", (2, 0), (2, 0), colors.HexColor("#CCFBF1")),
        ("BACKGROUND", (0, 1), (0, 1), colors.HexColor("#F0FDF4")),
        ("BACKGROUND", (1, 1), (1, 1), colors.HexColor("#FFF1F2")),
        ("BACKGROUND", (2, 1), (2, 1), colors.HexColor("#F0FDFA")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("FONTSIZE", (0, 1), (-1, 1), 13),
        ("ROUNDEDCORNERS", [6, 6, 6, 6]),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))

    elements.append(resumen_table)
    elements.append(Spacer(1, 0.3*inch))

    # Tabla de transacciones
    if transacciones:
        elements.append(Paragraph("Detalle de Transacciones", ParagraphStyle(
            "SectionTitle",
            parent=styles["Heading2"],
            fontSize=13,
            textColor=colors.HexColor("#0F172A"),
            spaceAfter=10
        )))

        headers = ["Fecha", "Descripción", "Categoría", "Cartera", "Tipo", "Monto"]
        tabla_data = [headers]

        for t in transacciones:
            tabla_data.append([
                t.get("fecha", ""),
                (t.get("descripcion", "") or "")[:30],
                (t.get("categories", {}) or {}).get("nombre", "")[:20],
                (t.get("wallets", {}) or {}).get("nombre", "")[:15],
                "Ingreso" if t.get("tipo") == "ingreso" else "Gasto",
                fmt(float(t.get("monto", 0)))
            ])

        col_widths = [0.9*inch, 1.8*inch, 1.3*inch, 1.0*inch, 0.8*inch, 1.1*inch]
        trans_table = Table(tabla_data, colWidths=col_widths, repeatRows=1)

        style = TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0D9488")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ALIGN", (1, 1), (1, -1), "LEFT"),
            ("ALIGN", (2, 1), (2, -1), "LEFT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#E2E8F0")),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ])

        # Colorear ingresos/gastos
        for i, t in enumerate(transacciones, start=1):
            if t.get("tipo") == "ingreso":
                style.add("TEXTCOLOR", (5, i), (5, i), colors.HexColor("#10B981"))
                style.add("FONTNAME", (5, i), (5, i), "Helvetica-Bold")
            else:
                style.add("TEXTCOLOR", (5, i), (5, i), colors.HexColor("#EF4444"))
                style.add("FONTNAME", (5, i), (5, i), "Helvetica-Bold")

        trans_table.setStyle(style)
        elements.append(trans_table)

    # Footer
    elements.append(Spacer(1, 0.3*inch))
    elements.append(Paragraph(
        f"Generado por Caudal — {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        ParagraphStyle("Footer", parent=styles["Normal"],
                      fontSize=8, textColor=colors.HexColor("#94A3B8"),
                      alignment=TA_CENTER)
    ))

    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=Caudal_{nombre_mes}.pdf"}
    )